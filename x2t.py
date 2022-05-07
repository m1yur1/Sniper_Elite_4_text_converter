import sys
import argparse
import io
from struct import pack
from openpyxl import load_workbook


class Record:
	name: str
	magic_number: str
	text: str


class AsrData:
	magic_number: str
	language_id: str
	file_name: str
	records: list


def read_xlsx(file_name):
	result = AsrData()
	result.records = []

	work_book = load_workbook(file_name)
	info_sheet = work_book['info']
	result.magic_number = info_sheet['A2'].value
	result.language_id = info_sheet['B2'].value
	result.file_name = info_sheet['C2'].value
	number_of_records = info_sheet['D2'].value

	data_sheet = work_book['data']
	for index in range(number_of_records):
		record = Record()
		record.magic_number = data_sheet.cell(index+1, 1).value
		record.name = data_sheet.cell(index+1, 2).value
		record.text = data_sheet.cell(index+1, 4).value
		if (None == record.text):
			record.text = ''
		result.records.append(record)

	return result


def unescape_text_length(text):
	escapes = text.count('\\')
	return len(text) - 4 * escapes + 1


def get_content_size(asr_data):
	content_bytes = 0
	text_bytes = 0
	name_bytes = 0

	for record in asr_data.records:
		text_bytes += 2 * unescape_text_length(record.text)
		name_bytes += len(text_to_ascii(record.name)) + 1

	# add: HTEXT, size, version, empty, records, magic, text_size, lang_id
	content_bytes += 4 * 8

	# add: record magic, record text_size
	content_bytes += 4 * 2 * len(asr_data.records)

	content_bytes += text_bytes + name_bytes
	file_name_bytes = len(asr_data.file_name) + 1

	padding_bytes = 4 - (file_name_bytes % 4)
	if (4 > padding_bytes):
		file_name_bytes += padding_bytes

	content_bytes += file_name_bytes

	# add: name_size
	content_bytes += 4

	return (content_bytes, text_bytes, name_bytes)


def text_to_utf16le(s):
	if not '\\' in s:
		return s.encode('utf_16_le')

	result = b''
	i = 0
	n = len(s)
	while True:
		if (n <= i):
			break

		if ('\\' != s[i]):
			result += s[i].encode('utf_16_le')
			i += 1
			continue

		v = s[i+1] + s[i+2]+ s[i+3] + s[i+4]
		i += 5
		result += pack('<H', int(v, 16))

	return result


def text_to_ascii(s):
	result = b''
	i = 0
	n = len(s)
	while True:
		if (n <= i):
			break

		if ('\\' != s[i]):
			result += s[i].encode('ascii')
			i += 1
			continue

		v = s[i+1] + s[i+2]
		i += 3
		result += pack('<B', int(v, 16))

	return result


def write_asr(file_name, asr_data):
	(content_bytes, text_bytes, name_bytes) = get_content_size(asr_data)
	with open(file_name, mode='wb') as f:
		f.write('Asura   '.encode('ascii'))
		f.write('HTXT'.encode('ascii'))
		f.write(pack('<I', content_bytes))
		f.write(pack('<I', 3))
		f.write(pack('<I', 0))
		f.write(pack('<I', len(asr_data.records)))
		f.write(pack('<I', int(asr_data.magic_number, 16)))
		f.write(pack('<I', text_bytes))
		f.write(pack('<I', int(asr_data.language_id, 16)))

		for record in asr_data.records:
			f.write(pack('<I', int(record.magic_number, 16)))
			f.write(pack('<I', unescape_text_length(record.text)))
			f.write(text_to_utf16le(record.text))
			f.write(pack('H', 0))

		f.write(asr_data.file_name.encode('ascii'))
		f.write(pack('B', 0))
		length = len(asr_data.file_name) + 1
		padd = 4 - (length % 4)
		if (4 == padd):
			padd = 0
			
		for _ in range(padd):
			f.write(pack('B', 0))

		f.write(pack('<I', name_bytes))
		for record in asr_data.records:
			f.write(text_to_ascii(record.name))
			f.write(pack('B', 0))

		for _ in range(16):
			f.write(pack('B', 0))
		
	return 0


def main(argv):
	parser = argparse.ArgumentParser(description='Convert *.xlsx to Sniper Elite 4 English text files (*.asr_en).')
	parser.add_argument('output_file_name', metavar='arg1', help='output *.asr_en file name.')
	parser.add_argument('input_file_name', metavar='arg2', help='input *.xlsx file name.')
	args = parser.parse_args()

	if (args.input_file_name.upper() == args.output_file_name.upper()):
		print('Input file name and output file name must be different.', file=sys.stderr)
		return -1

	d = read_xlsx(args.input_file_name)
	write_asr(args.output_file_name, d)


	return 0;


if '__main__' == __name__:
	sys.exit(main(sys.argv))
