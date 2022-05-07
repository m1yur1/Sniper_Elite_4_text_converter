import sys
import argparse
import io
from struct import unpack
from openpyxl import Workbook
from openpyxl.styles import numbers


class Record:
	name: str
	magic_number: str
	text: str


class AsrData:
	magic_number: str
	language_id: str
	file_name: str
	records: list


def utf16le_to_text(fragment_list):
	result = ''

	for fragment in fragment_list:
		us_data = unpack('<H', fragment)[0]
		if (0 == us_data):
			break

		#エスケープ不要の場合
		if (0x005c != us_data
			and (0x0020 <= us_data <= 0x007e
				or 0x3040 <= us_data <= 0x309F
				or 0x30A0 <= us_data <= 0x30FF
				or 0x3400 <= us_data <= 0x9FFF)):
			result += unpack('<2s', fragment)[0].decode('utf_16_le')
		#エスケープする場合
		else:
			result += '\\'
			result += str(format(us_data, '04x'))

	return result

def ascii_to_text(blob):
	result = ''

	for uc_data in blob:
		if (0 == uc_data):
			break
		#エスケープ不要の場合
		if (0x5c != uc_data
			and (0x20 <= uc_data <= 0x7e)):
			result += format(uc_data, 'c')
		#エスケープする場合
		else:
			result += '\\'
			result += format(uc_data, '02x')

	return result


def read_asr(file_name):
	result = AsrData()
	result.records = []

	with open(file_name, mode='rb') as f:
		# ヘッダ解析
		f.seek(8 + 4*4)
		items = unpack('<I', f.read(4))[0]
		result.magic_number = format(unpack('<I', f.read(4))[0], '08x')
		f.seek(4, io.SEEK_CUR)
		result.language_id = format(unpack('<I', f.read(4))[0], '08x')

		# 文字列の取得
		for _ in range(items):
			record = Record()
			record.name = ''
			record.magic_number = format(unpack('<I', f.read(4))[0], '08x')
			chars = unpack('<I', f.read(4))[0]

			fragment_list = []
			for _ in range(chars):
				fragment_list.append(f.read(2))

			record.text = utf16le_to_text(fragment_list)
			result.records.append(record)

		#ファイル名の取得
		file_name_in_asr = ''
		while True:
			fragment = unpack('<4c', f.read(4))
			for i in range(4):
				if (b'\x00' == fragment[i]):
					break
				else:
					file_name_in_asr += fragment[i].decode('ascii')

			if (b'\x00' == fragment[3]):
				break;

		result.file_name = file_name_in_asr

		# 文字列の名称の取得
		names_size = unpack('<I', f.read(4))[0]
		blob = f.read(names_size)
		names = blob.split(b'\x00')
		for i in range(len(result.records)):
			result.records[i].name = ascii_to_text(names[i])

	return result


def write_xlsx(file_name, asr_data):
	# 空のワークブックを用意する
	work_book = Workbook()
	while len(work_book.worksheets):
		work_book.remove(work_book.worksheets[-1])

	# infoワークシートを作り、セルを埋める
	info_sheet = work_book.create_sheet('info')
	info_sheet['A1'].value = 'magic_number'
	info_sheet['B1'].value = 'language_id'
	info_sheet['C1'].value = 'file_name'
	info_sheet['D1'].value = 'number_of_records'

	info_sheet['A2'].number_format = numbers.FORMAT_TEXT
	info_sheet['A2'].value = asr_data.magic_number

	info_sheet['B2'].number_format = numbers.FORMAT_TEXT
	info_sheet['B2'].value = asr_data.language_id

	info_sheet['C2'].number_format = numbers.FORMAT_TEXT
	info_sheet['C2'].value = asr_data.file_name

	info_sheet['D2'].number_format = numbers.FORMAT_NUMBER
	info_sheet['D2'].value = len(asr_data.records)

	# dataワークシートを作り、セルを埋める
	data_sheet = work_book.create_sheet('data')
	for index, record in enumerate(asr_data.records):
		data_sheet.cell(index+1, 1).number_format = numbers.FORMAT_TEXT
		data_sheet.cell(index+1, 1).value = record.magic_number

		data_sheet.cell(index+1, 2).number_format = numbers.FORMAT_TEXT
		data_sheet.cell(index+1, 2).value = record.name

		data_sheet.cell(index+1, 3).number_format = numbers.FORMAT_TEXT
		data_sheet.cell(index+1, 3).value = record.text

		data_sheet.cell(index+1, 4).number_format = numbers.FORMAT_TEXT
		data_sheet.cell(index+1, 4).value = record.text

	work_book.save(file_name)

	return 0


def main(argv):
	parser = argparse.ArgumentParser(description='Convert Sniper Elite 4 English text files (*.asr_en) to *.xlsx.')
	parser.add_argument('output_file_name', metavar='arg1', help='output *.xlsx file name.')
	parser.add_argument('input_file_name', metavar='arg2', help='input *.asr_en file name.')
	args = parser.parse_args()

	if (args.input_file_name.upper() == args.output_file_name.upper()):
		print('Input file name and output file name must be different.', file=sys.stderr)
		return -1

	d = read_asr(args.input_file_name)
	write_xlsx(args.output_file_name, d)

	return 0;


if '__main__' == __name__:
	sys.exit(main(sys.argv))
